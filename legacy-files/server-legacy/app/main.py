from __future__ import annotations

import hashlib
import json
from io import BytesIO
import os
import re
import sqlite3
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

from fastapi import Depends, FastAPI, HTTPException, UploadFile, File, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlmodel import Field, Session, SQLModel, create_engine, select
from uuid import uuid4

try:
    import firebase_admin
    from firebase_admin import credentials, messaging
except Exception:  # pragma: no cover
    firebase_admin = None
    credentials = None
    messaging = None

FCM_SERVICE_ACCOUNT = os.environ.get("FCM_SERVICE_ACCOUNT")
_fcm_app = None


def get_fcm_app():
    global _fcm_app
    if _fcm_app is not None:
        return _fcm_app
    if firebase_admin is None or not FCM_SERVICE_ACCOUNT:
        return None
    try:
        cred = credentials.Certificate(FCM_SERVICE_ACCOUNT)
        _fcm_app = firebase_admin.initialize_app(cred)
    except Exception:
        _fcm_app = None
    return _fcm_app


def send_push(tokens, title: str, body: str, data: Optional[dict] = None) -> None:
    if not tokens:
        return
    app = get_fcm_app()
    if app is None or messaging is None:
        return
    payload = {str(k): str(v) for k, v in (data or {}).items()}
    message = messaging.MulticastMessage(
        notification=messaging.Notification(title=title, body=body),
        data=payload,
        tokens=list(tokens),
    )
    try:
        messaging.send_multicast(message, app=app)
    except Exception:
        pass

BASE_DIR = Path(__file__).resolve().parent.parent
PROJECT_ROOT = BASE_DIR.parent
DATA_DIR = BASE_DIR / "data"
SCHEDULE_DIR = DATA_DIR / "schedule"
SCHEDULE_DIR.mkdir(parents=True, exist_ok=True)

NEWS_DIR = DATA_DIR / "news"
NEWS_DIR.mkdir(parents=True, exist_ok=True)

if str(PROJECT_ROOT) not in sys.path:
    sys.path.append(str(PROJECT_ROOT))

try:
    from alghoritm_schedule import generate_schedule_db
except Exception:
    generate_schedule_db = None

DB_PATH = BASE_DIR / "polyapp.db"
DB_URL = f"sqlite:///{DB_PATH.as_posix()}"

engine = create_engine(DB_URL, connect_args={"check_same_thread": False})

SECRET_KEY = os.environ.get("POLYAPP_SECRET", "dev-secret-change")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7
SESSION_TTL_DAYS = 7

pwd_context = CryptContext(schemes=["pbkdf2_sha256"], deprecated="auto")
security = HTTPBearer()

ROLE_LIST = [
    "smm",
    "parent",
    "request_handler",
    "admin",
    "student",
    "teacher",
]


REQUEST_TYPES = [
    "\u0421\u043f\u0440\u0430\u0432\u043a\u0430 \u043d\u0430 \u043e\u043d\u0430\u0439",
    "\u0421\u043f\u0440\u0430\u0432\u043a\u0430 \u043d\u0430 \u0432\u043e\u0435\u043d\u043a\u043e\u043c\u0430\u0442",
    "\u0421\u043f\u0440\u0430\u0432\u043a\u0430 \u043f\u043e \u043c\u0435\u0441\u0442\u0443 \u0442\u0440\u0435\u0431\u043e\u0432\u0430\u043d\u0438\u044f",
    "\u041f\u0440\u0438\u043b\u043e\u0436\u0435\u043d\u0438\u0435 \u21162",
    "\u041f\u0440\u0438\u043b\u043e\u0436\u0435\u043d\u0438\u0435 \u21164",
    "\u041f\u0440\u0438\u043b\u043e\u0436\u0435\u043d\u0438\u0435 \u21166",
    "\u041f\u0440\u0438\u043b\u043e\u0436\u0435\u043d\u0438\u0435 \u211629",
    "\u041f\u0440\u0438\u043b\u043e\u0436\u0435\u043d\u0438\u0435 \u211631",
    "\u0421\u043f\u0440\u0430\u0432\u043a\u0430 \u0432 \u0448\u043a\u043e\u043b\u0443",
]

REQUEST_STATUSES = [
    "\u041e\u0442\u043f\u0440\u0430\u0432\u043b\u0435\u043d\u0430",
    "\u041d\u0430 \u0440\u0430\u0441\u0441\u043c\u043e\u0442\u0440\u0435\u043d\u0438\u0438",
    "\u041e\u0442\u043a\u043b\u043e\u043d\u0435\u043d\u0430",
    "\u0412 \u0440\u0430\u0431\u043e\u0442\u0435",
    "\u0413\u043e\u0442\u043e\u0432\u0430",
]

class User(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    role: str
    full_name: str
    email: str
    password_hash: str
    phone: Optional[str] = None
    avatar_url: Optional[str] = None
    about: Optional[str] = None
    student_group: Optional[str] = None
    teacher_name: Optional[str] = None
    birth_date: Optional[date] = None


class UserPublic(SQLModel):
    id: int
    role: str
    full_name: str
    email: str
    phone: Optional[str] = None
    avatar_url: Optional[str] = None
    about: Optional[str] = None
    student_group: Optional[str] = None
    teacher_name: Optional[str] = None
    birth_date: Optional[date] = None


class NewsPost(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    title: str
    body: str
    author_id: int = Field(foreign_key="user.id")
    category: str = "news"
    pinned: bool = False
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: Optional[datetime] = None
    share_count: int = 0


class NewsMedia(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    post_id: int = Field(foreign_key="newspost.id")
    original_name: str
    stored_name: str
    media_type: str
    mime_type: Optional[str] = None
    size: Optional[int] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)


class NewsLike(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    post_id: int = Field(foreign_key="newspost.id")
    user_id: int = Field(foreign_key="user.id")
    reaction: str = "like"
    created_at: datetime = Field(default_factory=datetime.utcnow)


class NewsComment(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    post_id: int = Field(foreign_key="newspost.id")
    user_id: int = Field(foreign_key="user.id")
    text: str
    created_at: datetime = Field(default_factory=datetime.utcnow)


class ScheduleUpload(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    filename: str
    db_filename: Optional[str] = None
    schedule_date: Optional[date] = None
    uploaded_at: datetime = Field(default_factory=datetime.utcnow)


class RequestTicket(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    student_id: int = Field(foreign_key="user.id")
    request_type: str
    status: str = "\u041e\u0442\u043f\u0440\u0430\u0432\u043b\u0435\u043d\u0430"
    details: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)


class AttendanceRecord(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    group_name: str
    class_date: date
    student_name: str
    present: bool
    teacher_id: Optional[int] = Field(default=None, foreign_key="user.id")


class GradeRecord(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    group_name: str
    class_date: date
    student_name: str
    grade: int
    teacher_id: Optional[int] = Field(default=None, foreign_key="user.id")


class ExamGrade(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    group_name: str
    exam_name: str
    student_name: str
    grade: int
    created_at: datetime = Field(default_factory=datetime.utcnow)
    teacher_id: Optional[int] = Field(default=None, foreign_key="user.id")
    upload_id: Optional[int] = Field(default=None, foreign_key="examupload.id")


class ExamUpload(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    group_name: str
    exam_name: str
    filename: str
    rows_count: int
    uploaded_at: datetime = Field(default_factory=datetime.utcnow)
    teacher_id: Optional[int] = Field(default=None, foreign_key="user.id")


class TeacherGroupAssignment(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    teacher_id: int = Field(foreign_key="user.id")
    group_name: str
    subject: str
    created_at: datetime = Field(default_factory=datetime.utcnow)


class UserCreate(SQLModel):
    role: str
    full_name: str
    email: str
    password: str

class JournalGroup(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    name: str


class JournalStudent(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    group_name: str
    student_name: str


class JournalDate(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    group_name: str
    class_date: date



class UserUpdate(SQLModel):
    role: Optional[str] = None
    full_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    avatar_url: Optional[str] = None
    about: Optional[str] = None
    student_group: Optional[str] = None
    teacher_name: Optional[str] = None
    birth_date: Optional[date] = None


class NewsCreate(SQLModel):
    title: str
    body: str
    category: Optional[str] = None
    pinned: Optional[bool] = None


class NewsCommentCreate(SQLModel):
    text: str


class NewsUpdate(SQLModel):
    title: Optional[str] = None
    body: Optional[str] = None
    category: Optional[str] = None
    pinned: Optional[bool] = None


class NewsLikeAction(SQLModel):
    like: Optional[bool] = None
    reaction: Optional[str] = None


class NewsMediaOut(SQLModel):
    id: int
    url: str
    media_type: str
    original_name: str
    mime_type: Optional[str] = None
    size: Optional[int] = None


class NewsCommentOut(SQLModel):
    id: int
    user_id: int
    user_name: str
    text: str
    created_at: datetime


class NewsPostOut(SQLModel):
    id: int
    title: str
    body: str
    author_id: int
    author_name: str
    category: str
    pinned: bool
    created_at: datetime
    updated_at: Optional[datetime] = None
    share_count: int
    likes_count: int
    comments_count: int
    liked_by_me: bool
    reaction_counts: dict[str, int]
    my_reaction: Optional[str] = None
    media: list[NewsMediaOut]
    comments: list[NewsCommentOut]


class RequestTicketOut(SQLModel):
    id: int
    student_id: int
    student_name: str
    request_type: str
    status: str
    details: Optional[str] = None
    created_at: datetime


class ExamGradeOut(SQLModel):
    id: int
    group_name: str
    exam_name: str
    student_name: str
    grade: int
    created_at: datetime


class ExamUploadOut(SQLModel):
    id: int
    group_name: str
    exam_name: str
    filename: str
    rows_count: int
    uploaded_at: datetime
    teacher_name: Optional[str] = None


class ExamUploadUpdate(SQLModel):
    group_name: Optional[str] = None
    exam_name: Optional[str] = None


class TeacherGroupAssignmentOut(SQLModel):
    id: int
    teacher_id: int
    teacher_name: str
    group_name: str
    subject: str
    created_at: datetime


class TeacherGroupAssignmentCreate(SQLModel):
    teacher_id: int
    group_name: str
    subject: str


class TeacherGroupAssignmentUpdate(SQLModel):
    group_name: Optional[str] = None
    subject: Optional[str] = None


class GroupAnalytics(SQLModel):
    group_name: str
    subjects: list[str]
    teachers: list[str]


class RequestCreate(SQLModel):
    request_type: str
    details: Optional[str] = None


class RequestUpdate(SQLModel):
    status: Optional[str] = None
    details: Optional[str] = None


class AttendanceCreate(SQLModel):
    group_name: str
    class_date: date
    student_name: str
    present: bool


class GradeCreate(SQLModel):
    group_name: str
    class_date: date
    student_name: str
    grade: int


class GradeSummary(SQLModel):
    group_name: str
    average: float
    count: int

class JournalGroupCreate(SQLModel):
    name: str


class JournalStudentCreate(SQLModel):
    group_name: str
    student_name: str


class JournalDateCreate(SQLModel):
    group_name: str
    class_date: date



class AttendanceSummary(SQLModel):
    group_name: str
    present_count: int
    total_count: int


class ScheduleLesson(SQLModel):
    shift: int
    period: int
    time: str
    audience: str
    lesson: str
    group_name: str


class AuthRegister(SQLModel):
    role: Optional[str] = None
    full_name: str
    email: str
    password: str
    device_id: Optional[str] = None


class AuthLogin(SQLModel):
    email: str
    password: str
    device_id: Optional[str] = None


class TokenResponse(SQLModel):
    access_token: str
    token_type: str
    user: UserPublic


class AuthSession(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    session_id: str = Field(index=True, unique=True)
    user_id: int = Field(foreign_key="user.id")
    device_id: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    last_seen_at: datetime = Field(default_factory=datetime.utcnow)
    revoked_at: Optional[datetime] = None

class DeviceToken(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    user_id: int = Field(foreign_key="user.id")
    token: str = Field(index=True, unique=True)
    platform: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    last_seen_at: datetime = Field(default_factory=datetime.utcnow)
    revoked_at: Optional[datetime] = None


class Notification(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    user_id: int = Field(foreign_key="user.id")
    title: str
    body: str
    data_json: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    read_at: Optional[datetime] = None


class DeviceTokenRegister(SQLModel):
    token: str
    platform: Optional[str] = None


class NotificationOut(SQLModel):
    id: int
    title: str
    body: str
    data: Optional[dict] = None
    created_at: datetime
    read_at: Optional[datetime] = None




app = FastAPI(title="PolyApp API", version="0.2.0")

app.mount("/media/news", StaticFiles(directory=NEWS_DIR), name="news_media")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def ensure_user_schema() -> None:
    if not DB_PATH.exists():
        return
    conn = sqlite3.connect(DB_PATH.as_posix())
    try:
        cols = [row[1] for row in conn.execute("PRAGMA table_info(user)").fetchall()]
        if cols and 'birth_date' not in cols:
            conn.execute("ALTER TABLE user ADD COLUMN birth_date DATE")
            conn.commit()
    finally:
        conn.close()




def ensure_exam_schema() -> None:
    if not DB_PATH.exists():
        return
    conn = sqlite3.connect(DB_PATH.as_posix())
    try:
        cols = [row[1] for row in conn.execute("PRAGMA table_info(examgrade)").fetchall()]
        if cols and 'upload_id' not in cols:
            conn.execute("ALTER TABLE examgrade ADD COLUMN upload_id INTEGER")
        conn.commit()
    finally:
        conn.close()

def ensure_news_schema() -> None:
    if not DB_PATH.exists():
        return
    conn = sqlite3.connect(DB_PATH.as_posix())
    try:
        cols = [row[1] for row in conn.execute("PRAGMA table_info(newspost)").fetchall()]
        if cols:
            if 'updated_at' not in cols:
                conn.execute("ALTER TABLE newspost ADD COLUMN updated_at DATETIME")
            if 'share_count' not in cols:
                conn.execute("ALTER TABLE newspost ADD COLUMN share_count INTEGER NOT NULL DEFAULT 0")
            if 'category' not in cols:
                conn.execute("ALTER TABLE newspost ADD COLUMN category TEXT DEFAULT 'news'")
            if 'pinned' not in cols:
                conn.execute("ALTER TABLE newspost ADD COLUMN pinned INTEGER NOT NULL DEFAULT 0")
        like_cols = [row[1] for row in conn.execute("PRAGMA table_info(newslike)").fetchall()]
        if like_cols and 'reaction' not in like_cols:
            conn.execute("ALTER TABLE newslike ADD COLUMN reaction TEXT DEFAULT 'like'")
        conn.commit()
    finally:
        conn.close()


@app.on_event("startup")
def on_startup() -> None:
    SQLModel.metadata.create_all(engine)
    ensure_user_schema()
    ensure_news_schema()


@app.get("/health")
def health() -> dict:
    return {"status": "ok"}


@app.get("/roles")
def roles() -> list[str]:
    return ROLE_LIST


def to_public(user: User) -> UserPublic:
    return UserPublic(
        id=user.id,
        role=user.role,
        full_name=user.full_name,
        email=user.email,
        phone=user.phone,
        avatar_url=user.avatar_url,
        about=user.about,
        student_group=user.student_group,
        teacher_name=user.teacher_name,
        birth_date=user.birth_date,
    )


def hash_password(password: str) -> str:
    return pwd_context.hash(normalize_password(password))


def verify_password(password: str, password_hash: str) -> bool:
    return pwd_context.verify(normalize_password(password), password_hash)


def validate_password(password: str) -> None:
    if not password or not password.strip():
        raise HTTPException(status_code=400, detail="Password is required.")
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters.")
    if not re.search(r"[A-Z]", password):
        raise HTTPException(status_code=400, detail="Password must include an uppercase letter.")
    if not re.search(r"[a-z]", password):
        raise HTTPException(status_code=400, detail="Password must include a lowercase letter.")
    if not re.search(r"[0-9]", password):
        raise HTTPException(status_code=400, detail="Password must include a number.")


def normalize_password(password: str) -> str:
    raw = password.encode("utf-8")
    if len(raw) <= 72:
        return password
    return hashlib.sha256(raw).hexdigest()


def create_access_token(user: User, session_id: str) -> str:
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    payload = {"sub": str(user.id), "role": user.role, "sid": session_id, "exp": expire}
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(security),
) -> User:
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except JWTError as exc:
        raise HTTPException(status_code=401, detail="Invalid token") from exc
    user_id = payload.get("sub")
    session_id = payload.get("sid")
    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid token")
    if not session_id:
        raise HTTPException(status_code=401, detail="Invalid session")
    with Session(engine, expire_on_commit=False) as session:
        auth_session = session.exec(
            select(AuthSession).where(AuthSession.session_id == session_id)
        ).first()
        if not auth_session or auth_session.revoked_at is not None:
            raise HTTPException(status_code=401, detail="Session expired")
        if auth_session.last_seen_at < datetime.utcnow() - timedelta(days=SESSION_TTL_DAYS):
            auth_session.revoked_at = datetime.utcnow()
            session.add(auth_session)
            session.commit()
            raise HTTPException(status_code=401, detail="Session expired")
        user = session.get(User, int(user_id))
        if not user:
            raise HTTPException(status_code=401, detail="Invalid token")
        auth_session.last_seen_at = datetime.utcnow()
        session.add(auth_session)
        session.commit()
        return user




def create_notifications(
    session: Session,
    user_ids: list[int],
    title: str,
    body: str,
    data: Optional[dict] = None,
) -> None:
    if not user_ids:
        return
    payload = json.dumps(data) if data else None
    records = [
        Notification(user_id=user_id, title=title, body=body, data_json=payload)
        for user_id in user_ids
    ]
    session.add_all(records)
    session.commit()


def get_active_tokens(session: Session, user_ids: list[int]) -> list[str]:
    if not user_ids:
        return []
    tokens = session.exec(
        select(DeviceToken).where(
            DeviceToken.user_id.in_(user_ids),
            DeviceToken.revoked_at.is_(None),
        )
    ).all()
    return [t.token for t in tokens]

def require_roles(*roles: str):
    def dependency(current_user: User = Depends(get_current_user)) -> User:
        if current_user.role not in roles:
            raise HTTPException(status_code=403, detail="Insufficient role")
        return current_user

    return dependency


def parse_schedule_date(filename: str) -> Optional[date]:
    match = re.match(r"^schedule-(\d{2}\.\d{2}\.\d{4})\.sqlite$", filename)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%d.%m.%Y").date()
    except ValueError:
        return None


def sanitize_filename(name: str) -> str:
    base, dot, ext = name.rpartition('.')
    if dot == '':
        base = name
        ext = ''
    safe_base = re.sub(r'[^A-Za-z0-9._-]', '_', base)
    safe_ext = re.sub(r'[^A-Za-z0-9._-]', '_', ext)
    if not safe_base:
        safe_base = 'schedule'
    return safe_base + ('.' + safe_ext if safe_ext else '')


def classify_media(filename: str, content_type: Optional[str]) -> str:
    name = (filename or '').lower()
    if content_type and content_type.startswith('image/'):
        return 'image'
    if content_type and content_type.startswith('video/'):
        return 'video'
    if name.endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp')):
        return 'image'
    if name.endswith(('.mp4', '.mov', '.avi', '.mkv', '.webm')):
        return 'video'
    return 'document'


def extract_teacher_names(lesson_text: str) -> list[str]:
    if not lesson_text:
        return []
    candidates: set[str] = set()
    # Full names: ?????? ???? or ?????? ???? ????????
    for match in re.findall(r'[?-??][?-??]+\s+[?-??][?-??]+(?:\s+[?-??][?-??]+)?', lesson_text):
        candidates.add(match.strip())
    # Initials: ?????? ?.?. or ?????? ?. ?.
    for match in re.findall(r'[?-??][?-??]+\s+[?-??]\.?\s*[?-??]\.', lesson_text):
        candidates.add(match.strip())
    return sorted(candidates)


@app.post("/auth/register", response_model=TokenResponse)
def register(payload: AuthRegister) -> TokenResponse:
    role = (payload.role or "student").lower()
    if role not in ROLE_LIST:
        raise HTTPException(status_code=400, detail="Unknown role")
    validate_password(payload.password)
    with Session(engine) as session:
        existing = session.exec(select(User).where(User.email == payload.email)).first()
        if existing:
            raise HTTPException(status_code=409, detail="Email already registered")
        user = User(
            role=role,
            full_name=payload.full_name,
            email=payload.email,
            password_hash=hash_password(payload.password),
        )
        session.add(user)
        session.commit()
        session.refresh(user)
        session_id = uuid4().hex
        auth_session = AuthSession(
            session_id=session_id,
            user_id=user.id,
            device_id=payload.device_id,
        )
        session.add(auth_session)
        session.commit()
        token = create_access_token(user, session_id)
        return TokenResponse(access_token=token, token_type="bearer", user=to_public(user))


@app.post("/auth/login", response_model=TokenResponse)
def login(payload: AuthLogin) -> TokenResponse:
    validate_password(payload.password)
    with Session(engine) as session:
        user = session.exec(select(User).where(User.email == payload.email)).first()
        if not user or not verify_password(payload.password, user.password_hash):
            raise HTTPException(status_code=401, detail="Invalid credentials")
        session_id = uuid4().hex
        auth_session = AuthSession(
            session_id=session_id,
            user_id=user.id,
            device_id=payload.device_id,
        )
        session.add(auth_session)
        session.commit()
        token = create_access_token(user, session_id)
        return TokenResponse(access_token=token, token_type="bearer", user=to_public(user))


@app.get("/auth/me", response_model=UserPublic)
def me(current_user: User = Depends(get_current_user)) -> UserPublic:
    return to_public(current_user)


@app.post("/auth/logout")
def logout(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except JWTError as exc:
        raise HTTPException(status_code=401, detail="Invalid token") from exc
    session_id = payload.get("sid")
    if not session_id:
        raise HTTPException(status_code=401, detail="Invalid session")
    with Session(engine) as session:
        auth_session = session.exec(
            select(AuthSession).where(AuthSession.session_id == session_id)
        ).first()
        if auth_session and auth_session.revoked_at is None:
            auth_session.revoked_at = datetime.utcnow()
            session.add(auth_session)
            session.commit()
    return {"status": "ok"}




@app.post("/devices/register")
def register_device(
    payload: DeviceTokenRegister,
    current_user: User = Depends(get_current_user),
) -> dict:
    token = (payload.token or '').strip()
    if not token:
        raise HTTPException(status_code=400, detail="Missing token")
    with Session(engine) as session:
        existing = session.exec(select(DeviceToken).where(DeviceToken.token == token)).first()
        now = datetime.utcnow()
        if existing:
            existing.user_id = current_user.id
            existing.platform = payload.platform
            existing.last_seen_at = now
            existing.revoked_at = None
            session.add(existing)
        else:
            session.add(
                DeviceToken(
                    user_id=current_user.id,
                    token=token,
                    platform=payload.platform,
                    created_at=now,
                    last_seen_at=now,
                )
            )
        session.commit()
    return {"status": "ok"}


@app.get("/notifications", response_model=list[NotificationOut])
def list_notifications(
    offset: int = 0,
    limit: int = 50,
    current_user: User = Depends(get_current_user),
) -> list[NotificationOut]:
    with Session(engine) as session:
        records = session.exec(
            select(Notification)
            .where(Notification.user_id == current_user.id)
            .order_by(Notification.created_at.desc())
            .offset(offset)
            .limit(limit)
        ).all()
        result: list[NotificationOut] = []
        for item in records:
            data = None
            if item.data_json:
                try:
                    data = json.loads(item.data_json)
                except Exception:
                    data = None
            result.append(
                NotificationOut(
                    id=item.id,
                    title=item.title,
                    body=item.body,
                    data=data,
                    created_at=item.created_at,
                    read_at=item.read_at,
                )
            )
        return result


@app.post("/notifications/{notification_id}/read", response_model=NotificationOut)
def mark_notification_read(
    notification_id: int,
    current_user: User = Depends(get_current_user),
) -> NotificationOut:
    with Session(engine) as session:
        notif = session.get(Notification, notification_id)
        if not notif or notif.user_id != current_user.id:
            raise HTTPException(status_code=404, detail="Notification not found")
        if notif.read_at is None:
            notif.read_at = datetime.utcnow()
            session.add(notif)
            session.commit()
        data = None
        if notif.data_json:
            try:
                data = json.loads(notif.data_json)
            except Exception:
                data = None
        return NotificationOut(
            id=notif.id,
            title=notif.title,
            body=notif.body,
            data=data,
            created_at=notif.created_at,
            read_at=notif.read_at,
        )

@app.post("/users", response_model=UserPublic)
def create_user(
    payload: UserCreate,
    current_user: User = Depends(require_roles("admin")),
) -> UserPublic:
    if payload.role not in ROLE_LIST:
        raise HTTPException(status_code=400, detail="Unknown role")
    with Session(engine) as session:
        existing = session.exec(select(User).where(User.email == payload.email)).first()
        if existing:
            raise HTTPException(status_code=409, detail="Email already registered")
        user = User(
            role=role,
            full_name=payload.full_name,
            email=payload.email,
            password_hash=hash_password(payload.password),
        )
        session.add(user)
        session.commit()
        session.refresh(user)
        return to_public(user)


@app.get("/users", response_model=list[UserPublic])
def list_users(
    role: Optional[str] = None,
    current_user: User = Depends(require_roles("admin")),
) -> list[UserPublic]:
    with Session(engine) as session:
        statement = select(User)
        if role:
            statement = statement.where(User.role == role)
        users = session.exec(statement).all()
        return [to_public(user) for user in users]



@app.get("/users/{user_id}", response_model=UserPublic)
def get_user(user_id: int, current_user: User = Depends(get_current_user)) -> UserPublic:
    if current_user.role != "admin" and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Insufficient role")
    with Session(engine) as session:
        user = session.get(User, user_id)
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        return to_public(user)


@app.patch("/users/{user_id}", response_model=UserPublic)
def update_user(
    user_id: int,
    payload: UserUpdate,
    current_user: User = Depends(get_current_user),
) -> UserPublic:
    if current_user.role != "admin" and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Insufficient role")
    with Session(engine) as session:
        user = session.get(User, user_id)
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        updates = payload.dict(exclude_unset=True)
        if current_user.role != "admin":
            updates.pop("role", None)
        for key, value in updates.items():
            setattr(user, key, value)
        session.add(user)
        session.commit()
        session.refresh(user)
        return to_public(user)


def build_news_post_out(post: NewsPost, current_user: User, session: Session) -> NewsPostOut:
    author = session.get(User, post.author_id)
    media_items = list(session.exec(select(NewsMedia).where(NewsMedia.post_id == post.id)).all())
    comments = list(session.exec(
        select(NewsComment).where(NewsComment.post_id == post.id).order_by(NewsComment.created_at.asc())
    ).all())
    likes = list(session.exec(select(NewsLike).where(NewsLike.post_id == post.id)).all())
    my_like = session.exec(
        select(NewsLike).where(NewsLike.post_id == post.id, NewsLike.user_id == current_user.id)
    ).first()
    liked_by_me = my_like is not None

    user_names: dict[int, str] = {}
    for comment in comments:
        if comment.user_id not in user_names:
            user = session.get(User, comment.user_id)
            user_names[comment.user_id] = user.full_name if user else 'Unknown'

    reaction_counts: dict[str, int] = {}
    for like in likes:
        reaction = like.reaction or 'like'
        reaction_counts[reaction] = reaction_counts.get(reaction, 0) + 1

    return NewsPostOut(
        id=post.id,
        title=post.title,
        body=post.body,
        author_id=post.author_id,
        author_name=author.full_name if author else 'Unknown',
        category=post.category or 'news',
        pinned=bool(post.pinned),
        created_at=post.created_at,
        updated_at=post.updated_at,
        share_count=post.share_count,
        likes_count=len(likes),
        comments_count=len(comments),
        liked_by_me=liked_by_me,
        reaction_counts=reaction_counts,
        my_reaction=my_like.reaction if my_like else None,
        media=[
            NewsMediaOut(
                id=item.id,
                url=f"/media/news/{item.stored_name}",
                media_type=item.media_type,
                original_name=item.original_name,
                mime_type=item.mime_type,
                size=item.size,
            )
            for item in media_items
        ],
        comments=[
            NewsCommentOut(
                id=comment.id,
                user_id=comment.user_id,
                user_name=user_names.get(comment.user_id, 'Unknown'),
                text=comment.text,
                created_at=comment.created_at,
            )
            for comment in comments
        ],
    )


@app.post("/news", response_model=NewsPostOut)
async def create_news(
    request: Request,
    files: list[UploadFile] = File(default=[]),
    title: Optional[str] = Form(default=None),
    body: Optional[str] = Form(default=None),
    category: Optional[str] = Form(default=None),
    pinned: Optional[bool] = Form(default=None),
    current_user: User = Depends(require_roles("smm", "admin")),
) -> NewsPostOut:
    content_type = request.headers.get('content-type', '')
    if content_type.startswith('application/json'):
        data = await request.json()
        title = data.get('title')
        body = data.get('body')
        category = data.get('category')
        pinned = data.get('pinned')
    title = (title or '').strip()
    body = (body or '').strip()
    if isinstance(pinned, str):
        pinned = pinned.lower() in ('true', '1', 'yes')
    if not title and not body and not files:
        raise HTTPException(status_code=400, detail='Empty post')

    post = NewsPost(
        title=title,
        body=body,
        author_id=current_user.id,
        category=(category or "news"),
        pinned=bool(pinned) if pinned is not None else False,
    )
    with Session(engine) as session:
        session.add(post)
        session.commit()
        session.refresh(post)
        for file in files:
            if not file.filename:
                continue
            suffix = Path(file.filename).suffix
            stored_name = f"{uuid4().hex}{suffix.lower()}"
            target = NEWS_DIR / stored_name
            content = file.file.read()
            with target.open('wb') as buffer:
                buffer.write(content)
            media = NewsMedia(
                post_id=post.id,
                original_name=file.filename,
                stored_name=stored_name,
                media_type=classify_media(file.filename, file.content_type),
                mime_type=file.content_type,
                size=len(content) if content is not None else None,
            )
            session.add(media)
        post.updated_at = datetime.utcnow()
        session.add(post)
        session.commit()
        session.refresh(post)
        return build_news_post_out(post, current_user, session)


@app.get("/news", response_model=list[NewsPostOut])
def list_news(
    current_user: User = Depends(get_current_user),
    offset: int = 0,
    limit: int = 20,
    category: Optional[str] = None,
) -> list[NewsPostOut]:
    limit = max(1, min(limit, 50))
    offset = max(0, offset)
    with Session(engine) as session:
        statement = select(NewsPost)
        if category:
            statement = statement.where(NewsPost.category == category)
        posts = list(
            session.exec(
                statement
                .order_by(NewsPost.pinned.desc(), NewsPost.created_at.desc())
                .offset(offset)
                .limit(limit)
            ).all()
        )
        return [build_news_post_out(post, current_user, session) for post in posts]


@app.get("/news/{post_id}", response_model=NewsPostOut)
def get_news_post(
    post_id: int,
    current_user: User = Depends(get_current_user),
) -> NewsPostOut:
    with Session(engine) as session:
        post = session.get(NewsPost, post_id)
        if not post:
            raise HTTPException(status_code=404, detail='Post not found')
        return build_news_post_out(post, current_user, session)



@app.patch("/news/{post_id}", response_model=NewsPostOut)
def update_news(
    post_id: int,
    payload: NewsUpdate,
    current_user: User = Depends(get_current_user),
) -> NewsPostOut:
    if current_user.role not in ("smm", "admin"):
        raise HTTPException(status_code=403, detail="Insufficient role")
    with Session(engine) as session:
        post = session.get(NewsPost, post_id)
        if not post:
            raise HTTPException(status_code=404, detail="Post not found")
        if payload.title is not None:
            post.title = payload.title
        if payload.body is not None:
            post.body = payload.body
        if payload.category is not None:
            post.category = payload.category
        if payload.pinned is not None:
            post.pinned = payload.pinned
        post.updated_at = datetime.utcnow()
        session.add(post)
        session.commit()
        session.refresh(post)
        return build_news_post_out(post, current_user, session)


@app.post("/news/{post_id}/like")
def like_news(
    post_id: int,
    payload: NewsLikeAction | None = None,
    current_user: User = Depends(get_current_user),
) -> dict:
    with Session(engine) as session:
        post = session.get(NewsPost, post_id)
        if not post:
            raise HTTPException(status_code=404, detail='Post not found')
        existing = session.exec(
            select(NewsLike).where(NewsLike.post_id == post_id, NewsLike.user_id == current_user.id)
        ).first()
        desired = payload.like if payload is not None else None
        reaction = (payload.reaction if payload is not None else None) or 'like'
        if desired is False:
            if existing is not None:
                session.delete(existing)
                session.commit()
        else:
            if existing is None:
                session.add(NewsLike(post_id=post_id, user_id=current_user.id, reaction=reaction))
                session.commit()
            else:
                if existing.reaction != reaction:
                    existing.reaction = reaction
                    session.add(existing)
                    session.commit()
        likes = list(session.exec(select(NewsLike).where(NewsLike.post_id == post_id)).all())
        liked = session.exec(
            select(NewsLike).where(NewsLike.post_id == post_id, NewsLike.user_id == current_user.id)
        ).first()
        reaction_counts: dict[str, int] = {}
        for like in likes:
            r = like.reaction or 'like'
            reaction_counts[r] = reaction_counts.get(r, 0) + 1
        return {
            "likes_count": len(likes),
            "liked": liked is not None,
            "reaction_counts": reaction_counts,
            "my_reaction": liked.reaction if liked else None,
        }


@app.post("/news/{post_id}/comment", response_model=NewsCommentOut)
def comment_news(
    post_id: int,
    payload: NewsCommentCreate,
    current_user: User = Depends(get_current_user),
) -> NewsCommentOut:
    text = payload.text.strip()
    if not text:
        raise HTTPException(status_code=400, detail='Empty comment')
    with Session(engine) as session:
        post = session.get(NewsPost, post_id)
        if not post:
            raise HTTPException(status_code=404, detail='Post not found')
        comment = NewsComment(post_id=post_id, user_id=current_user.id, text=text)
        session.add(comment)
        post.updated_at = datetime.utcnow()
        session.add(post)
        session.commit()
        session.refresh(comment)
        return NewsCommentOut(
            id=comment.id,
            user_id=comment.user_id,
            user_name=current_user.full_name,
            text=comment.text,
            created_at=comment.created_at,
        )


@app.delete("/news/{post_id}")
def delete_news(
    post_id: int,
    current_user: User = Depends(get_current_user),
) -> dict:
    if current_user.role not in ("smm", "admin"):
        raise HTTPException(status_code=403, detail="Insufficient role")
    with Session(engine) as session:
        post = session.get(NewsPost, post_id)
        if not post:
            raise HTTPException(status_code=404, detail="Post not found")
        media_items = list(session.exec(select(NewsMedia).where(NewsMedia.post_id == post_id)).all())
        for media in media_items:
            session.delete(media)
            file_path = NEWS_DIR / media.stored_name
            try:
                if file_path.exists():
                    file_path.unlink()
            except Exception:
                pass
        for comment in session.exec(select(NewsComment).where(NewsComment.post_id == post_id)).all():
            session.delete(comment)
        for like in session.exec(select(NewsLike).where(NewsLike.post_id == post_id)).all():
            session.delete(like)
        session.delete(post)
        session.commit()
    return {"status": "ok"}


@app.delete("/news/{post_id}/comment/{comment_id}")
def delete_news_comment(
    post_id: int,
    comment_id: int,
    current_user: User = Depends(get_current_user),
) -> dict:
    if current_user.role not in ("smm", "admin"):
        raise HTTPException(status_code=403, detail="Insufficient role")
    with Session(engine) as session:
        comment = session.get(NewsComment, comment_id)
        if not comment or comment.post_id != post_id:
            raise HTTPException(status_code=404, detail="Comment not found")
        session.delete(comment)
        session.commit()
    return {"status": "ok"}


@app.post("/news/{post_id}/share")
def share_news(
    post_id: int,
    current_user: User = Depends(get_current_user),
) -> dict:
    with Session(engine) as session:
        post = session.get(NewsPost, post_id)
        if not post:
            raise HTTPException(status_code=404, detail='Post not found')
        post.share_count = (post.share_count or 0) + 1
        post.updated_at = datetime.utcnow()
        session.add(post)
        session.commit()
        return {"share_count": post.share_count}


@app.post("/schedule/upload", response_model=ScheduleUpload)
def upload_schedule(
    file: UploadFile,
    current_user: User = Depends(require_roles("admin")),
) -> ScheduleUpload:
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing filename")
    safe_name = sanitize_filename(file.filename)
    destination = SCHEDULE_DIR / safe_name
    with destination.open("wb") as buffer:
        buffer.write(file.file.read())

    db_filename = None
    schedule_date = None
    if generate_schedule_db is None:
        raise HTTPException(status_code=500, detail="Schedule parser not available")
    db_path = Path(generate_schedule_db(str(destination), str(SCHEDULE_DIR)))
    db_filename = db_path.name
    schedule_date = parse_schedule_date(db_filename)

    record = ScheduleUpload(
        filename=safe_name,
        db_filename=db_filename,
        schedule_date=schedule_date,
    )
    with Session(engine) as session:
        session.add(record)
        session.commit()
        session.refresh(record)
        title = "Расписание обновлено"
        body = "Доступно новое расписание"
        data = {"type": "schedule", "date": str(schedule_date) if schedule_date else ""}
        roles = ["admin", "teacher", "student"]
        users = session.exec(select(User).where(User.role.in_(roles))).all()
        user_ids = [u.id for u in users]
        create_notifications(session, user_ids, title, body, data)
        tokens = get_active_tokens(session, user_ids)
        send_push(tokens, title, body, data)
        return record


@app.get("/schedule", response_model=list[ScheduleUpload])
def list_schedule(current_user: User = Depends(get_current_user)) -> list[ScheduleUpload]:
    with Session(engine) as session:
        return list(session.exec(select(ScheduleUpload).order_by(ScheduleUpload.uploaded_at.desc())).all())


@app.get("/schedule/latest", response_model=Optional[ScheduleUpload])
def get_latest_schedule(current_user: User = Depends(get_current_user)) -> Optional[ScheduleUpload]:
    with Session(engine) as session:
        return session.exec(select(ScheduleUpload).order_by(ScheduleUpload.uploaded_at.desc())).first()


@app.get("/schedule/groups", response_model=list[str])
def list_schedule_groups(current_user: User = Depends(get_current_user)) -> list[str]:
    with Session(engine) as session:
        latest = session.exec(select(ScheduleUpload).order_by(ScheduleUpload.uploaded_at.desc())).first()
        if not latest or not latest.db_filename:
            return []
    db_path = SCHEDULE_DIR / latest.db_filename
    if not db_path.exists():
        return []
    conn = sqlite3.connect(db_path.as_posix())
    try:
        rows = conn.execute("SELECT DISTINCT group_name FROM lessons ORDER BY group_name").fetchall()
    finally:
        conn.close()
    return [row[0] for row in rows if row[0]]


@app.get("/schedule/teachers", response_model=list[str])
def list_schedule_teachers(current_user: User = Depends(get_current_user)) -> list[str]:
    with Session(engine) as session:
        latest = session.exec(select(ScheduleUpload).order_by(ScheduleUpload.uploaded_at.desc())).first()
        if not latest or not latest.db_filename:
            return []
    db_path = SCHEDULE_DIR / latest.db_filename
    if not db_path.exists():
        return []
    conn = sqlite3.connect(db_path.as_posix())
    names: set[str] = set()
    try:
        for (lesson,) in conn.execute("SELECT lesson FROM lessons").fetchall():
            for name in extract_teacher_names(lesson or ''):
                names.add(name)
    finally:
        conn.close()
    return sorted(names)


@app.get("/schedule/me", response_model=list[ScheduleLesson])
def schedule_for_me(current_user: User = Depends(get_current_user)) -> list[ScheduleLesson]:
    with Session(engine) as session:
        latest = session.exec(select(ScheduleUpload).order_by(ScheduleUpload.uploaded_at.desc())).first()
        if not latest or not latest.db_filename:
            return []
    db_path = SCHEDULE_DIR / latest.db_filename
    if not db_path.exists():
        return []
    conn = sqlite3.connect(db_path.as_posix())
    try:
        if current_user.role == "student":
            if not current_user.student_group:
                return []
            rows = conn.execute(
                """
                SELECT shift, period, time, audience, lesson, group_name
                FROM lessons
                WHERE group_name = ?
                ORDER BY shift ASC, period ASC, audience ASC
                """
                , (current_user.student_group,)
            ).fetchall()
        elif current_user.role == "teacher":
            if not current_user.teacher_name:
                return []
            tokens = [t.strip() for t in current_user.teacher_name.split() if t.strip()]
            if not tokens:
                return []
            rows = conn.execute(
                """
                SELECT shift, period, time, audience, lesson, group_name
                FROM lessons
                ORDER BY shift ASC, period ASC, audience ASC
                """
            ).fetchall()
            lowered_tokens = [t.lower() for t in tokens]
            rows = [
                row for row in rows
                if all(token in (row[4] or '').lower() for token in lowered_tokens)
            ]
        else:
            return []
    finally:
        conn.close()
    return [
        ScheduleLesson(
            shift=row[0],
            period=row[1],
            time=row[2],
            audience=row[3],
            lesson=row[4],
            group_name=row[5],
        )
        for row in rows
    ]


@app.get("/schedule/teacher/{teacher_name}", response_model=list[ScheduleLesson])
def schedule_for_teacher(
    teacher_name: str,
    current_user: User = Depends(require_roles("admin")),
) -> list[ScheduleLesson]:
    with Session(engine) as session:
        latest = session.exec(select(ScheduleUpload).order_by(ScheduleUpload.uploaded_at.desc())).first()
        if not latest or not latest.db_filename:
            return []
    db_path = SCHEDULE_DIR / latest.db_filename
    if not db_path.exists():
        return []
    tokens = [t.strip() for t in teacher_name.split() if t.strip()]
    if not tokens:
        return []
    conn = sqlite3.connect(db_path.as_posix())
    try:
        rows = conn.execute(
            """
            SELECT shift, period, time, audience, lesson, group_name
            FROM lessons
            ORDER BY shift ASC, period ASC, audience ASC
            """
        ).fetchall()
    finally:
        conn.close()
    lowered_tokens = [t.lower() for t in tokens]
    rows = [row for row in rows if all(token in (row[4] or '').lower() for token in lowered_tokens)]
    return [
        ScheduleLesson(
            shift=row[0],
            period=row[1],
            time=row[2],
            audience=row[3],
            lesson=row[4],
            group_name=row[5],
        )
        for row in rows
    ]


@app.get("/schedule/group/{group_name}", response_model=list[ScheduleLesson])
def schedule_for_group(
    group_name: str,
    current_user: User = Depends(require_roles("admin", "student", "teacher")),
) -> list[ScheduleLesson]:
    with Session(engine) as session:
        latest = session.exec(select(ScheduleUpload).order_by(ScheduleUpload.uploaded_at.desc())).first()
        if not latest or not latest.db_filename:
            return []
    db_path = SCHEDULE_DIR / latest.db_filename
    if not db_path.exists():
        raise HTTPException(status_code=404, detail="Schedule database not found")

    conn = sqlite3.connect(db_path.as_posix())
    try:
        rows = conn.execute(
            """
            SELECT shift, period, time, audience, lesson, group_name
            FROM lessons
            WHERE group_name = ?
            ORDER BY shift ASC, period ASC, audience ASC
            """,
            (group_name,),
        ).fetchall()
    finally:
        conn.close()

    return [
        ScheduleLesson(
            shift=row[0],
            period=row[1],
            time=row[2],
            audience=row[3],
            lesson=row[4],
            group_name=row[5],
        )
        for row in rows
    ]




@app.get("/journal/groups", response_model=list[str])
def list_journal_groups(current_user: User = Depends(require_roles("teacher", "admin", "parent"))) -> list[str]:
    with Session(engine) as session:
        groups = session.exec(select(JournalGroup).order_by(JournalGroup.name.asc())).all()
        return [g.name for g in groups]


@app.post("/journal/groups", response_model=JournalGroup)
def upsert_journal_group(
    payload: JournalGroupCreate,
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> JournalGroup:
    with Session(engine) as session:
        existing = session.exec(select(JournalGroup).where(JournalGroup.name == payload.name)).first()
        if existing:
            return existing
        group = JournalGroup(name=payload.name)
        session.add(group)
        session.commit()
        session.refresh(group)
        return group


@app.delete("/journal/groups")
def delete_journal_group(
    group_name: str,
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> dict:
    with Session(engine) as session:
        group = session.exec(select(JournalGroup).where(JournalGroup.name == group_name)).first()
        if group:
            session.delete(group)
        # cascade deletes in related tables
        students = session.exec(select(JournalStudent).where(JournalStudent.group_name == group_name)).all()
        for st in students:
            session.delete(st)
        dates = session.exec(select(JournalDate).where(JournalDate.group_name == group_name)).all()
        for dt in dates:
            session.delete(dt)
        attendance = session.exec(select(AttendanceRecord).where(AttendanceRecord.group_name == group_name)).all()
        for at in attendance:
            session.delete(at)
        grades = session.exec(select(GradeRecord).where(GradeRecord.group_name == group_name)).all()
        for gr in grades:
            session.delete(gr)
        session.commit()
    return {"status": "ok"}


@app.get("/journal/students", response_model=list[str])
def list_journal_students(
    group_name: str,
    current_user: User = Depends(require_roles("teacher", "admin", "parent")),
) -> list[str]:
    with Session(engine) as session:
        students = session.exec(
            select(JournalStudent).where(JournalStudent.group_name == group_name).order_by(JournalStudent.student_name.asc())
        ).all()
        return [s.student_name for s in students]


@app.post("/journal/students", response_model=JournalStudent)
def upsert_journal_student(
    payload: JournalStudentCreate,
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> JournalStudent:
    with Session(engine) as session:
        existing = session.exec(
            select(JournalStudent).where(
                JournalStudent.group_name == payload.group_name,
                JournalStudent.student_name == payload.student_name,
            )
        ).first()
        if existing:
            return existing
        student = JournalStudent(group_name=payload.group_name, student_name=payload.student_name)
        session.add(student)
        session.commit()
        session.refresh(student)
        return student


@app.delete("/journal/students")
def delete_journal_student(
    group_name: str,
    student_name: str,
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> dict:
    with Session(engine) as session:
        student = session.exec(
            select(JournalStudent).where(
                JournalStudent.group_name == group_name,
                JournalStudent.student_name == student_name,
            )
        ).first()
        if student:
            session.delete(student)
        attendance = session.exec(
            select(AttendanceRecord).where(
                AttendanceRecord.group_name == group_name,
                AttendanceRecord.student_name == student_name,
            )
        ).all()
        for at in attendance:
            session.delete(at)
        grades = session.exec(
            select(GradeRecord).where(
                GradeRecord.group_name == group_name,
                GradeRecord.student_name == student_name,
            )
        ).all()
        for gr in grades:
            session.delete(gr)
        session.commit()
    return {"status": "ok"}


@app.get("/journal/dates", response_model=list[date])
def list_journal_dates(
    group_name: str,
    current_user: User = Depends(require_roles("teacher", "admin", "parent")),
) -> list[date]:
    with Session(engine) as session:
        dates = session.exec(
            select(JournalDate).where(JournalDate.group_name == group_name).order_by(JournalDate.class_date.asc())
        ).all()
        return [d.class_date for d in dates]


@app.post("/journal/dates", response_model=JournalDate)
def upsert_journal_date(
    payload: JournalDateCreate,
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> JournalDate:
    with Session(engine) as session:
        existing = session.exec(
            select(JournalDate).where(
                JournalDate.group_name == payload.group_name,
                JournalDate.class_date == payload.class_date,
            )
        ).first()
        if existing:
            return existing
        date_record = JournalDate(group_name=payload.group_name, class_date=payload.class_date)
        session.add(date_record)
        session.commit()
        session.refresh(date_record)
        return date_record


@app.delete("/journal/dates")
def delete_journal_date(
    group_name: str,
    class_date: date,
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> dict:
    with Session(engine) as session:
        date_record = session.exec(
            select(JournalDate).where(
                JournalDate.group_name == group_name,
                JournalDate.class_date == class_date,
            )
        ).first()
        if date_record:
            session.delete(date_record)
        attendance = session.exec(
            select(AttendanceRecord).where(
                AttendanceRecord.group_name == group_name,
                AttendanceRecord.class_date == class_date,
            )
        ).all()
        for at in attendance:
            session.delete(at)
        grades = session.exec(
            select(GradeRecord).where(
                GradeRecord.group_name == group_name,
                GradeRecord.class_date == class_date,
            )
        ).all()
        for gr in grades:
            session.delete(gr)
        session.commit()
    return {"status": "ok"}


@app.post("/requests", response_model=RequestTicketOut)
def create_request(
    payload: RequestCreate,
    current_user: User = Depends(require_roles("student")),
) -> RequestTicketOut:
    req_type = (payload.request_type or '').strip().lower()
    if req_type not in REQUEST_TYPES:
        raise HTTPException(status_code=400, detail="Unknown request type")
    if not current_user.full_name.strip() or not current_user.phone or not current_user.birth_date or not current_user.student_group:
        raise HTTPException(status_code=400, detail="Profile incomplete")
    ticket = RequestTicket(
        student_id=current_user.id,
        request_type=req_type,
        details=payload.details,
    )
    with Session(engine) as session:
        session.add(ticket)
        session.commit()
        session.refresh(ticket)
        return RequestTicketOut(
            id=ticket.id,
            student_id=ticket.student_id,
            student_name=current_user.full_name,
            request_type=ticket.request_type,
            status=ticket.status,
            details=ticket.details,
            created_at=ticket.created_at,
        )


@app.get("/requests", response_model=list[RequestTicketOut])
def list_requests(
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user),
) -> list[RequestTicketOut]:
    with Session(engine) as session:
        statement = select(RequestTicket)
        if status:
            statement = statement.where(RequestTicket.status == status)
        if current_user.role == "student":
            statement = statement.where(RequestTicket.student_id == current_user.id)
        tickets = list(session.exec(statement.order_by(RequestTicket.created_at.desc())).all())
        user_ids = {t.student_id for t in tickets}
        users = session.exec(select(User).where(User.id.in_(user_ids))).all() if user_ids else []
        name_map = {u.id: u.full_name for u in users}
        return [
            RequestTicketOut(
                id=t.id,
                student_id=t.student_id,
                student_name=name_map.get(t.student_id, 'Unknown'),
                request_type=t.request_type,
                status=t.status,
                details=t.details,
                created_at=t.created_at,
            )
            for t in tickets
        ]


@app.delete("/requests/{ticket_id}")
def delete_request(
    ticket_id: int,
    current_user: User = Depends(require_roles("request_handler", "admin")),
) -> dict:
    with Session(engine) as session:
        ticket = session.get(RequestTicket, ticket_id)
        if not ticket:
            raise HTTPException(status_code=404, detail="Request not found")
        session.delete(ticket)
        session.commit()
    return {"status": "ok"}


@app.patch("/requests/{ticket_id}", response_model=RequestTicketOut)
def update_request(
    ticket_id: int,
    payload: RequestUpdate,
    current_user: User = Depends(require_roles("request_handler", "admin")),
) -> RequestTicketOut:
    with Session(engine) as session:
        ticket = session.get(RequestTicket, ticket_id)
        if not ticket:
            raise HTTPException(status_code=404, detail="Request not found")
        if payload.status is not None:
            if payload.status not in REQUEST_STATUSES:
                raise HTTPException(status_code=400, detail="Unknown status")
            ticket.status = payload.status
        if payload.details is not None:
            ticket.details = payload.details
        session.add(ticket)
        session.commit()
        session.refresh(ticket)
        student = session.get(User, ticket.student_id)
        return RequestTicketOut(
            id=ticket.id,
            student_id=ticket.student_id,
            student_name=student.full_name if student else 'Unknown',
            request_type=ticket.request_type,
            status=ticket.status,
            details=ticket.details,
            created_at=ticket.created_at,
        )


@app.get("/teacher-assignments", response_model=list[TeacherGroupAssignmentOut])
def list_teacher_assignments(
    group_name: Optional[str] = None,
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> list[TeacherGroupAssignmentOut]:
    with Session(engine) as session:
        statement = select(TeacherGroupAssignment)
        if current_user.role == "teacher":
            statement = statement.where(TeacherGroupAssignment.teacher_id == current_user.id)
        if group_name:
            statement = statement.where(TeacherGroupAssignment.group_name == group_name)
        assignments = list(session.exec(statement.order_by(TeacherGroupAssignment.group_name.asc())).all())
        teacher_ids = {a.teacher_id for a in assignments}
        teachers = session.exec(select(User).where(User.id.in_(teacher_ids))).all() if teacher_ids else []
        name_map = {u.id: u.full_name for u in teachers}
        return [
            TeacherGroupAssignmentOut(
                id=a.id,
                teacher_id=a.teacher_id,
                teacher_name=name_map.get(a.teacher_id, "Unknown"),
                group_name=a.group_name,
                subject=a.subject,
                created_at=a.created_at,
            )
            for a in assignments
        ]


@app.post("/teacher-assignments", response_model=TeacherGroupAssignmentOut)
def create_teacher_assignment(
    payload: TeacherGroupAssignmentCreate,
    current_user: User = Depends(require_roles("admin")),
) -> TeacherGroupAssignmentOut:
    if not payload.group_name.strip() or not payload.subject.strip():
        raise HTTPException(status_code=400, detail="Group and subject are required")
    with Session(engine) as session:
        teacher = session.get(User, payload.teacher_id)
        if not teacher or teacher.role != "teacher":
            raise HTTPException(status_code=404, detail="Teacher not found")
        assignment = TeacherGroupAssignment(
            teacher_id=payload.teacher_id,
            group_name=payload.group_name.strip(),
            subject=payload.subject.strip(),
        )
        session.add(assignment)
        session.commit()
        session.refresh(assignment)
        return TeacherGroupAssignmentOut(
            id=assignment.id,
            teacher_id=assignment.teacher_id,
            teacher_name=teacher.full_name,
            group_name=assignment.group_name,
            subject=assignment.subject,
            created_at=assignment.created_at,
        )


@app.patch("/teacher-assignments/{assignment_id}", response_model=TeacherGroupAssignmentOut)
def update_teacher_assignment(
    assignment_id: int,
    payload: TeacherGroupAssignmentUpdate,
    current_user: User = Depends(require_roles("admin")),
) -> TeacherGroupAssignmentOut:
    with Session(engine) as session:
        assignment = session.get(TeacherGroupAssignment, assignment_id)
        if not assignment:
            raise HTTPException(status_code=404, detail="Assignment not found")
        if payload.group_name is not None:
            assignment.group_name = payload.group_name
        if payload.subject is not None:
            assignment.subject = payload.subject
        session.add(assignment)
        session.commit()
        session.refresh(assignment)
        teacher = session.get(User, assignment.teacher_id)
        return TeacherGroupAssignmentOut(
            id=assignment.id,
            teacher_id=assignment.teacher_id,
            teacher_name=teacher.full_name if teacher else "Unknown",
            group_name=assignment.group_name,
            subject=assignment.subject,
            created_at=assignment.created_at,
        )


@app.delete("/teacher-assignments/{assignment_id}")
def delete_teacher_assignment(
    assignment_id: int,
    current_user: User = Depends(require_roles("admin")),
) -> dict:
    with Session(engine) as session:
        assignment = session.get(TeacherGroupAssignment, assignment_id)
        if not assignment:
            raise HTTPException(status_code=404, detail="Assignment not found")
        session.delete(assignment)
        session.commit()
        return {"status": "ok"}


@app.get("/analytics/groups", response_model=list[GroupAnalytics])
def analytics_groups(
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> list[GroupAnalytics]:
    with Session(engine) as session:
        assignments = session.exec(select(TeacherGroupAssignment)).all()
        if current_user.role == "teacher":
            assignments = [a for a in assignments if a.teacher_id == current_user.id]
        group_names = {a.group_name for a in assignments}
        # include groups from journal for admin
        if current_user.role == "admin":
            groups = session.exec(select(JournalGroup)).all()
            for g in groups:
                group_names.add(g.name)
        # map teacher names
        teacher_ids = {a.teacher_id for a in assignments}
        teachers = session.exec(select(User).where(User.id.in_(teacher_ids))).all() if teacher_ids else []
        name_map = {u.id: u.full_name for u in teachers}
        result = []
        for name in sorted(group_names):
            subjects = sorted({a.subject for a in assignments if a.group_name == name})
            teacher_names = sorted({name_map.get(a.teacher_id, "Unknown") for a in assignments if a.group_name == name})
            result.append(GroupAnalytics(group_name=name, subjects=subjects, teachers=teacher_names))
        return result


@app.get("/analytics/attendance", response_model=list[AttendanceRecord])
def analytics_attendance(
    group_name: Optional[str] = None,
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> list[AttendanceRecord]:
    with Session(engine) as session:
        allowed_groups: Optional[set[str]] = None
        if current_user.role == "teacher":
            allowed_groups = {a.group_name for a in session.exec(select(TeacherGroupAssignment).where(TeacherGroupAssignment.teacher_id == current_user.id)).all()}
        statement = select(AttendanceRecord)
        if group_name:
            statement = statement.where(AttendanceRecord.group_name == group_name)
        records = list(session.exec(statement.order_by(AttendanceRecord.class_date.desc())).all())
        if allowed_groups is not None:
            records = [r for r in records if r.group_name in allowed_groups]
        return records


@app.get("/analytics/grades", response_model=list[GradeRecord])
def analytics_grades(
    group_name: Optional[str] = None,
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> list[GradeRecord]:
    with Session(engine) as session:
        allowed_groups: Optional[set[str]] = None
        if current_user.role == "teacher":
            allowed_groups = {a.group_name for a in session.exec(select(TeacherGroupAssignment).where(TeacherGroupAssignment.teacher_id == current_user.id)).all()}
        statement = select(GradeRecord)
        if group_name:
            statement = statement.where(GradeRecord.group_name == group_name)
        records = list(session.exec(statement.order_by(GradeRecord.class_date.desc())).all())
        if allowed_groups is not None:
            records = [r for r in records if r.group_name in allowed_groups]
        return records


@app.get("/exams", response_model=list[ExamGradeOut])
def list_exam_grades(
    group_name: Optional[str] = None,
    exam_name: Optional[str] = None,
    current_user: User = Depends(require_roles("student", "parent", "teacher", "admin")),
) -> list[ExamGradeOut]:
    with Session(engine) as session:
        statement = select(ExamGrade)
        if current_user.role == "student":
            statement = statement.where(ExamGrade.student_name == current_user.full_name)
            if current_user.student_group:
                statement = statement.where(ExamGrade.group_name == current_user.student_group)
        elif current_user.role == "parent":
            if current_user.student_group:
                statement = statement.where(ExamGrade.group_name == current_user.student_group)
        if group_name:
            statement = statement.where(ExamGrade.group_name == group_name)
        if exam_name:
            statement = statement.where(ExamGrade.exam_name == exam_name)
        records = list(session.exec(statement.order_by(ExamGrade.created_at.desc())).all())
        return [
            ExamGradeOut(
                id=r.id,
                group_name=r.group_name,
                exam_name=r.exam_name,
                student_name=r.student_name,
                grade=r.grade,
                created_at=r.created_at,
            )
            for r in records
        ]


@app.get("/exams/uploads", response_model=list[ExamUploadOut])
def list_exam_uploads(
    group_name: Optional[str] = None,
    exam_name: Optional[str] = None,
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> list[ExamUploadOut]:
    with Session(engine) as session:
        statement = select(ExamUpload)
        if current_user.role == "teacher":
            statement = statement.where(ExamUpload.teacher_id == current_user.id)
        if group_name:
            statement = statement.where(ExamUpload.group_name == group_name)
        if exam_name:
            statement = statement.where(ExamUpload.exam_name == exam_name)
        uploads = list(session.exec(statement.order_by(ExamUpload.uploaded_at.desc())).all())
        teacher_ids = {u.teacher_id for u in uploads if u.teacher_id}
        teachers = session.exec(select(User).where(User.id.in_(teacher_ids))).all() if teacher_ids else []
        name_map = {u.id: u.full_name for u in teachers}
        return [
            ExamUploadOut(
                id=u.id,
                group_name=u.group_name,
                exam_name=u.exam_name,
                filename=u.filename,
                rows_count=u.rows_count,
                uploaded_at=u.uploaded_at,
                teacher_name=name_map.get(u.teacher_id),
            )
            for u in uploads
        ]


@app.patch("/exams/uploads/{upload_id}", response_model=ExamUploadOut)
def update_exam_upload(
    upload_id: int,
    payload: ExamUploadUpdate,
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> ExamUploadOut:
    if payload.group_name is None and payload.exam_name is None:
        raise HTTPException(status_code=400, detail="Nothing to update")
    with Session(engine) as session:
        upload = session.get(ExamUpload, upload_id)
        if not upload:
            raise HTTPException(status_code=404, detail="Upload not found")
        if current_user.role == "teacher" and upload.teacher_id != current_user.id:
            raise HTTPException(status_code=403, detail="Forbidden")
        if payload.group_name is not None:
            upload.group_name = payload.group_name
        if payload.exam_name is not None:
            upload.exam_name = payload.exam_name
        session.add(upload)
        grades = session.exec(select(ExamGrade).where(ExamGrade.upload_id == upload_id)).all()
        for grade in grades:
            if payload.group_name is not None:
                grade.group_name = payload.group_name
            if payload.exam_name is not None:
                grade.exam_name = payload.exam_name
            session.add(grade)
        session.commit()
        session.refresh(upload)
        teacher = session.get(User, upload.teacher_id) if upload.teacher_id else None
        return ExamUploadOut(
            id=upload.id,
            group_name=upload.group_name,
            exam_name=upload.exam_name,
            filename=upload.filename,
            rows_count=upload.rows_count,
            uploaded_at=upload.uploaded_at,
            teacher_name=teacher.full_name if teacher else None,
        )


@app.delete("/exams/uploads/{upload_id}")
def delete_exam_upload(
    upload_id: int,
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> dict:
    with Session(engine) as session:
        upload = session.get(ExamUpload, upload_id)
        if not upload:
            raise HTTPException(status_code=404, detail="Upload not found")
        if current_user.role == "teacher" and upload.teacher_id != current_user.id:
            raise HTTPException(status_code=403, detail="Forbidden")
        grades = session.exec(select(ExamGrade).where(ExamGrade.upload_id == upload_id)).all()
        for grade in grades:
            session.delete(grade)
        session.delete(upload)
        session.commit()
        return {"deleted": len(grades)}


@app.post("/exams/upload", response_model=list[ExamGradeOut])
def upload_exam_grades(
    group_name: str = Form(...),
    exam_name: str = Form(...),
    file: UploadFile = File(...),
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> list[ExamGradeOut]:
    if not group_name.strip() or not exam_name.strip():
        raise HTTPException(status_code=400, detail="Group and exam name are required")
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing file")
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(status_code=500, detail="Excel parser not available") from exc
    data = file.file.read()
    wb = load_workbook(BytesIO(data), data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        name = row[0] if len(row) > 0 else None
        grade_val = row[1] if len(row) > 1 else None
        if name is None:
            continue
        student_name = str(name).strip()
        if not student_name:
            continue
        if grade_val is None:
            continue
        if isinstance(grade_val, (int, float)):
            grade = int(grade_val)
        else:
            try:
                grade = int(str(grade_val).strip())
            except ValueError:
                raise HTTPException(status_code=400, detail=f"Invalid grade for {student_name}")
        rows.append((student_name, grade))

    if not rows:
        raise HTTPException(status_code=400, detail="No grades found")

    with Session(engine) as session:
        upload = ExamUpload(
            group_name=group_name.strip(),
            exam_name=exam_name.strip(),
            filename=file.filename,
            rows_count=len(rows),
            teacher_id=current_user.id,
        )
        session.add(upload)
        session.flush()
        records = []
        for student_name, grade in rows:
            record = ExamGrade(
                group_name=group_name.strip(),
                exam_name=exam_name.strip(),
                student_name=student_name,
                grade=grade,
                teacher_id=current_user.id,
                upload_id=upload.id,
            )
            session.add(record)
            records.append(record)
        session.commit()
        for record in records:
            session.refresh(record)
        student_names = {name for name, _ in rows}
        users = session.exec(select(User).where(User.role.in_(["student", "parent"])) ).all()
        user_ids: list[int] = []
        for user in users:
            if user.role == "student":
                if user.full_name in student_names and (user.student_group is None or user.student_group == group_name.strip()):
                    user_ids.append(user.id)
            elif user.role == "parent":
                if user.student_group and user.student_group == group_name.strip():
                    user_ids.append(user.id)
        title = f"Новые экзаменационные оценки: {exam_name.strip()}"
        body = f"Группа {group_name.strip()}"
        data = {"type": "exam_grades", "group": group_name.strip(), "exam": exam_name.strip()}
        create_notifications(session, user_ids, title, body, data)
        tokens = get_active_tokens(session, user_ids)
        send_push(tokens, title, body, data)
        return [
            ExamGradeOut(
                id=r.id,
                group_name=r.group_name,
                exam_name=r.exam_name,
                student_name=r.student_name,
                grade=r.grade,
                created_at=r.created_at,
            )
            for r in records
        ]


@app.post("/attendance", response_model=AttendanceRecord)
def create_attendance(
    payload: AttendanceCreate,
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> AttendanceRecord:
    with Session(engine) as session:
        existing = session.exec(
            select(AttendanceRecord).where(
                AttendanceRecord.group_name == payload.group_name,
                AttendanceRecord.class_date == payload.class_date,
                AttendanceRecord.student_name == payload.student_name,
            )
        ).first()
        if existing:
            existing.present = payload.present
            existing.teacher_id = current_user.id
            session.add(existing)
            session.commit()
            session.refresh(existing)
            return existing
        record = AttendanceRecord(
            group_name=payload.group_name,
            class_date=payload.class_date,
            student_name=payload.student_name,
            present=payload.present,
            teacher_id=current_user.id,
        )
        session.add(record)
        session.commit()
        session.refresh(record)
        return record


@app.get("/attendance", response_model=list[AttendanceRecord])
def list_attendance(
    group_name: Optional[str] = None,
    current_user: User = Depends(require_roles("teacher", "admin", "parent")),
) -> list[AttendanceRecord]:
    with Session(engine) as session:
        statement = select(AttendanceRecord)
        if group_name:
            statement = statement.where(AttendanceRecord.group_name == group_name)
        return list(session.exec(statement.order_by(AttendanceRecord.class_date.desc())).all())




@app.delete("/attendance")
def delete_attendance(
    group_name: str,
    class_date: date,
    student_name: str,
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> dict:
    with Session(engine) as session:
        statement = select(AttendanceRecord).where(
            AttendanceRecord.group_name == group_name,
            AttendanceRecord.class_date == class_date,
            AttendanceRecord.student_name == student_name,
        )
        records = session.exec(statement).all()
        for record in records:
            session.delete(record)
        session.commit()
        return {"deleted": len(records)}


@app.get("/attendance/summary", response_model=list[AttendanceSummary])
def attendance_summary(
    group_name: Optional[str] = None,
    current_user: User = Depends(require_roles("teacher", "admin", "parent")),
) -> list[AttendanceSummary]:
    with Session(engine) as session:
        statement = select(AttendanceRecord)
        if group_name:
            statement = statement.where(AttendanceRecord.group_name == group_name)
        records = session.exec(statement).all()
    totals: dict[str, AttendanceSummary] = {}
    for record in records:
        entry = totals.get(
            record.group_name,
            AttendanceSummary(group_name=record.group_name, present_count=0, total_count=0),
        )
        entry.total_count += 1
        if record.present:
            entry.present_count += 1
        totals[record.group_name] = entry
    return list(totals.values())


@app.post("/grades", response_model=GradeRecord)
def create_grade(
    payload: GradeCreate,
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> GradeRecord:
    if payload.grade < 1 or payload.grade > 100:
        raise HTTPException(status_code=400, detail="Grade must be 1..100")
    with Session(engine) as session:
        existing = session.exec(
            select(GradeRecord).where(
                GradeRecord.group_name == payload.group_name,
                GradeRecord.class_date == payload.class_date,
                GradeRecord.student_name == payload.student_name,
            )
        ).first()
        if existing:
            existing.grade = payload.grade
            existing.teacher_id = current_user.id
            session.add(existing)
            session.commit()
            session.refresh(existing)
            return existing
        record = GradeRecord(
            group_name=payload.group_name,
            class_date=payload.class_date,
            student_name=payload.student_name,
            grade=payload.grade,
            teacher_id=current_user.id,
        )
        session.add(record)
        session.commit()
        session.refresh(record)
        return record


@app.get("/grades", response_model=list[GradeRecord])
def list_grades(
    group_name: Optional[str] = None,
    current_user: User = Depends(require_roles("teacher", "admin", "parent")),
) -> list[GradeRecord]:
    with Session(engine) as session:
        statement = select(GradeRecord)
        if group_name:
            statement = statement.where(GradeRecord.group_name == group_name)
        return list(session.exec(statement.order_by(GradeRecord.class_date.desc())).all())




@app.delete("/grades")
def delete_grade(
    group_name: str,
    class_date: date,
    student_name: str,
    current_user: User = Depends(require_roles("teacher", "admin")),
) -> dict:
    with Session(engine) as session:
        statement = select(GradeRecord).where(
            GradeRecord.group_name == group_name,
            GradeRecord.class_date == class_date,
            GradeRecord.student_name == student_name,
        )
        records = session.exec(statement).all()
        for record in records:
            session.delete(record)
        session.commit()
        return {"deleted": len(records)}


@app.get("/grades/summary", response_model=list[GradeSummary])
def grade_summary(
    group_name: Optional[str] = None,
    current_user: User = Depends(require_roles("teacher", "admin", "parent")),
) -> list[GradeSummary]:
    with Session(engine) as session:
        statement = select(GradeRecord)
        if group_name:
            statement = statement.where(GradeRecord.group_name == group_name)
        records = session.exec(statement).all()
    totals: dict[str, tuple[int, int]] = {}
    for record in records:
        total, count = totals.get(record.group_name, (0, 0))
        totals[record.group_name] = (total + record.grade, count + 1)
    summaries: list[GradeSummary] = []
    for group, (total, count) in totals.items():
        average = round(total / count, 2) if count else 0.0
        summaries.append(GradeSummary(group_name=group, average=average, count=count))
    return summaries
